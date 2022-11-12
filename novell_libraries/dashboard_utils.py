# -*- coding: utf-8 -*-
# coding=utf-8
from __future__ import unicode_literals

import ast
import calendar
import csv
import datetime
import hashlib
import json
import logging
import shutil
import socket
import string
import sys
import tarfile
import uuid
from ConfigParser import ConfigParser
from hashlib import md5

import codecs
import os
import paramiko
import re
import xlsxwriter
from algorithms.sorting import bubble_sort
from openpyxl import load_workbook
from sqlalchemy import and_
from sqlalchemy.orm.exc import NoResultFound
from sqlalchemy.sql.expression import between

"""
    :author: Jesus Becerril Navarrete
    :organization: NOVELL MICROFOCUS
    :contact: jesus.becerril@microfocus.com

"""

__docformat__ = "restructuredtext"


months = {1: 'Enero', 2: 'Febrero', 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
          7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"}
reload(sys)

sys.setdefaultencoding('utf8')


class Logs:
    '''Clase destinada a los objetos que manipulan los archivos de texto o logs de esta aplicacion'''

    def __init__(self, file_log):
        """
        Funcion que inicializa la clase es muy similar a un constructor da de alta los objetos de clase logger para
        poder usarlos para escribir los mensajes en los archivos de log
        :param file_log: Archivo donde se escriran los mensajes de log.
        """
        self.logger = None
        if None is self.logger:
            self.logger = logging.getLogger('dashboard_SLA')
            self.hdlr = logging.FileHandler(file_log)
            formatter = logging.Formatter('%(asctime)s %(levelname)s %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')
            self.hdlr.setFormatter(formatter)
            self.logger.addHandler(self.hdlr)
            self.logger.setLevel(logging.DEBUG)

    def info_log(self, message):
        """
        Funcion usada para crear el log con el nivel de prioridades de nivel 6  o mensajes informativos
        :param message: string que contiene el mensaje a escribir en el log
        """
        self.logger.info(message)

    #        self.logger.removeHandler(self.hdlr)

    def error_log(self, message):
        """
        Funcion usada para crear el log con el nivel de prioridades de nivel 3  o mensajes de error
        :param message: string que contiene el mensaje a escribir en el log
        """
        self.logger.error(message)


class Utils:
    def __init__(self):
        pass

    @staticmethod
    def identity_validator(name):
        identity_hash = hashlib.sha256(uuid.uuid5(uuid.NAMESPACE_DNS, str(name)).get_hex()).hexdigest()
        return identity_hash

    @staticmethod
    def createMd5sum(files):
        chunk = 8192
        hasher = md5()
        try:
            afile = open(files, 'rb')
            buf = afile.read(chunk)
            while len(buf) > 0:
                hasher.update(buf)
                buf = afile.read(chunk)
            return hasher.hexdigest()

        except:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            print "Se precento el error: %s %s %s0" % (exc_type, exc_value, exc_traceback)


class FileConfig:
    def __init__(self):
        pass

    @staticmethod
    def db_config():
        """
        Funcion que devuelve la intefaz con la que estara trabajando la aplicacion la cual esta configurada en el
        archivo de texto plano usado para las configuracione.
        :returns:ifacenet: Devuelve la interfaz con la que se trabajara en ek equipo
        """
        file_config = "/etc/dashboard/dashboard.conf"
        config = ConfigParser()
        config.read(file_config)
        user = config.get('credentials', 'db_user')
        password = config.get('credentials', 'db_password')
        server = config.get('credentials', 'server_db')
        return user, password, server

    @staticmethod
    def key_location():
        file_config = "/etc/dashboard/dashboard.conf"
        config = ConfigParser()
        config.read(file_config)
        key = config.get('ssh_data', 'key_file')
        user = config.get('ssh_data', 'user')
        return key, user

    @staticmethod
    def google_data():
        file_config = "/etc/dashboard/dashboard.conf"
        config = ConfigParser()
        config.read(file_config)
        spreadsheets_changes = config.get('google_data', 'google_spreadsheet_id')
        spreadsheets_incidents = config.get('google_data', 'google_spreadsheet_in')
        file_cred = config.get('google_data', 'google_creds_file')
        file_key = config.get('google_data', 'google_file_key')
        proxy_data = config.get('google_data', 'proxy')
        return spreadsheets_changes, file_key, file_cred, proxy_data, spreadsheets_incidents

    @staticmethod
    def inventory_file():
        file_config = "/etc/dashboard/dashboard.conf"
        config = ConfigParser()
        config.read(file_config)
        inventory_file = config.get('files', 'inventory_file')

        return inventory_file

    @staticmethod
    def reciver_file():
        file_config = "/etc/dashboard/dashboard.conf"
        config = ConfigParser()
        config.read(file_config)
        reciver_file = config.get('files', 'reciver_file')
        return reciver_file

    @staticmethod
    def local_receiver_file():
        file_config = "/etc/dashboard/dashboard.conf"
        config = ConfigParser()
        config.read(file_config)
        receiver_file = config.get('files', 'local_receiver')
        return receiver_file

    @staticmethod
    def google_file():
        file_config = "/etc/dashboard/dashboard.conf"
        config = ConfigParser()
        config.read(file_config)
        google_file = config.get('files', 'google_file')
        return google_file

    @staticmethod
    def error_conn_file():
        file_config = "/etc/dashboard/dashboard.conf"
        config = ConfigParser()
        config.read(file_config)
        error_file = config.get('files', 'error_file')
        return error_file

    @staticmethod
    def user_log_file():
        file_config = "/etc/dashboard/dashboard.conf"
        config = ConfigParser()
        config.read(file_config)
        user_log = config.get('files', 'user_log')
        return user_log

    @staticmethod
    def cmd():
        file_config = "/etc/dashboard/dashboard.conf"
        config = ConfigParser()
        config.read(file_config)
        cmd_ruby = config.get('cmd', 'cmd_ruby')
        return cmd_ruby

    @staticmethod
    def certs():
        file_config = "/etc/dashboard/dashboard.conf"
        config = ConfigParser()
        config.read(file_config)
        key = config.get('ssl_certs', 'KEY_CERT')
        cert = config.get('ssl_certs', 'FILE_CERT')
        return cert, key

    @staticmethod
    def menus():
        file_config = "/etc/dashboard/dashboard.conf"
        f = codecs.open(file_config, 'r', encoding='utf-8')
        config = ConfigParser()
        config.readfp(f)
        platform = ast.literal_eval(config.get('menus', 'platform'))
        environment = ast.literal_eval(config.get('menus', 'environment'))
        status = ast.literal_eval(config.get('menus', 'status'))
        support_group = ast.literal_eval(config.get('menus', 'support_group'))
        priority = ast.literal_eval(config.get('menus', 'priority'))
        vobo = ast.literal_eval(config.get('menus', 'vobo'))
        category = ast.literal_eval(config.get('menus', 'category'))

        return list(platform), list(environment), list(status), list(support_group), list(priority), list(vobo), list(
            category)

    @staticmethod
    def environment_data():
        file_config = "/etc/dashboard/dashboard.conf"
        config = ConfigParser()
        config.read(file_config)
        debug = config.get('environment_data', 'DEBUG')
        debug_sql = config.get('environment_data', 'SQLALCHEMY_ECHO')
        secret_key = config.get('environment_data', 'SECRET_KEY')
        session = config.get('environment_data', 'SESSION_PROTECTION')
        dir_upload = config.get('environment_data', 'UPLOAD_FOLDER')
        extensions = config.get('environment_data', 'EXTENSIONS')
        sla_log_store = config.get('environment_data', 'STORE')
        tmp_log_process = config.get('environment_data', 'SLA_LOG_TMP')
        environment_dict = ast.literal_eval(config.get('environment_data', 'ENVIRONMENT_DICT'))

        return debug, secret_key, session, dir_upload, extensions, sla_log_store, tmp_log_process, debug_sql, \
               environment_dict


class SshCon:
    def __init__(self):
        pass

    @staticmethod
    def ssh_conn(server_ip, logs, ini="", fin=""):
        try:
            data_conn = FileConfig().key_location()
            c = paramiko.SSHClient()
            c._policy = paramiko.WarningPolicy()
            rsa_key = paramiko.RSAKey.from_private_key_file(data_conn[0])
            c.connect(server_ip, username=data_conn[1], pkey=rsa_key)
            cmd = "/usr/bin/server_reports  " + ini + "  " + fin
            logs.info_log(u"Ejecutando server_reports")
            stdin, stdout, stderr = c.exec_command(cmd)
            s_inf = stdout.read()
            if s_inf is not None and s_inf != "":
                j_inf = json.dumps(str(s_inf).splitlines()[1])
                c.close()
                logs.info_log(u"Server reports se ejecuto sin problemas")
                return j_inf
        except:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            logs.error_log(u"Se presento el error: %s %s %s0" % (exc_type, exc_value, exc_traceback))


class UploadInfoFile:
    def __init__(self):
        pass

    def get_data(self, ws, row):
        list_data = []
        column_list = list(string.uppercase[:8])
        while True:
            try:
                data_tuple = ws[column_list[0] + str(row)].value, ws[column_list[1] + str(row)].value, \
                             ws[column_list[2] + str(row)].value, ws[column_list[3] + str(row)].value, \
                             ws[column_list[4] + str(row)].value, ws[column_list[5] + str(row)].value, \
                             ws[column_list[6] + str(row)].value, ws[column_list[7] + str(row)].value,

                if data_tuple is not None:
                    list_data.append(data_tuple)
                else:
                    break
                #
                row += 1
            except Exception:
                exc_type, exc_value, exc_traceback = sys.exc_info()
                #                print "Se presento el error: %s %s %s0" % (exc_type, exc_value, exc_traceback)
                break

        return list_data

    def upload_info(self, db, file_name, servers):
        wb = load_workbook(filename=file_name, read_only=True)
        sheet_name = wb.get_sheet_names()[0]
        ws = wb.get_sheet_by_name(sheet_name)
        aux = 2
        data = self.get_data(ws, aux)
        if len(data) > 0:
            utils = Utils()
            for data_server in data:
                bio = utils.identity_validator(data_server[0])
                try:

                    if data_server[0] != None and data_server[0] != "":
                        q = db.session.query(servers).filter(servers.hostname == data_server[0]).one()
                        q.hostname = data_server[0].lower()
                        q.date_admission = str(datetime.datetime.now().strftime("%Y/%m/%d-%H:%M"))

                        if data_server[1].lower().strip() == "none":
                            try:
                                q.ip_admin = str(socket.gethostbyname(data_server[0].lower().strip()))
                            except:
                                q.ip_admin = data_server[1].lower()

                        q.biometric = bio
                        q.ip_prod = data_server[3].lower()
                        q.environment = data_server[2].lower()
                        q.status = data_server[7].lower()
                        q.service = data_server[6]
                        q.platform = data_server[4]
                        q.service_group = data_server[5]
                        db.session.commit()

                except NoResultFound:
                    ip_ad = 'none'
                    if data_server[1].lower() == "none":
                        try:
                            ip_ad = socket.gethostbyname(data_server[0].lower().strip())
                        except:
                            ip_ad = data_server[1].lower()

                    server = servers(hostname=data_server[0].lower(),
                                     date_admission=str(datetime.datetime.now().strftime("%Y/%m/%d-%H:%M")),
                                     ip_admin=ip_ad,
                                     biometric=bio, ip_prod=data_server[3].lower(),
                                     environment=data_server[2].lower(),
                                     status=data_server[7].lower(),
                                     service=data_server[6],
                                     platform=data_server[4],
                                     service_group=data_server[5])

                    db.session.add(server)
                    db.session.commit()

    def upload_compress(self, db, file_name, store):
        HOME = os.getcwd()
        environment_dict = FileConfig().environment_data()[8]
        n_regex = re.compile('[0-9]')
        s_regex = re.compile('[a-z]')
        name_f = file_name.split("/")[2]
        env = name_f.split("-")[0]
        month = n_regex.split(name_f.split("-")[1])[0]
        year = s_regex.split(name_f.split(".")[0].split("-")[1])[-1]
        log_store = FileConfig().environment_data()[5] + file_name.split("/")[2].split("-")[1].split(".")[0]
        store_file_name = log_store + "/" + name_f
        if os.path.exists(log_store) and os.path.isdir(log_store):
            if not os.path.exists(store_file_name):
                shutil.move(file_name, log_store)
            elif os.path.exists(store_file_name) and os.path.isfile(store_file_name):
                md5_s = Utils().createMd5sum(store_file_name)
                md5_tmp = Utils().createMd5sum(file_name)
                if md5_s == md5_tmp:
                    pass
                else:
                    shutil.move(file_name, log_store)
        else:
            os.makedirs(log_store)
            if not os.path.exists(store_file_name):
                shutil.move(file_name, log_store)
            elif os.path.exists(store_file_name) and os.path.isfile(store_file_name):
                md5_s = Utils().createMd5sum(store_file_name)
                md5_tmp = Utils().createMd5sum(file_name)
                if md5_s == md5_tmp:
                    pass
                else:
                    shutil.move(file_name, log_store)

        md5 = Utils().createMd5sum(store_file_name)
        try:
            q = db.session.query(store).filter(and_(store.file_name == name_f, store.hash_md5 == md5)).one()
            os.chdir(HOME)
        except NoResultFound:
            sla_logs = store(environment=environment_dict[env], absolute_route=store_file_name, file_name=name_f,
                             hash_md5=md5, month=month, year=year.strip("."))
            db.session.add(sla_logs)
            db.session.commit()
            os.chdir(HOME)

    def extract_log(self, store_file, store_md5, store_fname):
        HOME = os.getcwd()
        try:
            proc_dir = FileConfig().environment_data()[6]
            dir_log = store_fname.split(".")[0] + "/"
            if os.path.exists(proc_dir) and os.path.isdir(proc_dir):
                if os.path.exists(proc_dir + store_fname) and os.path.isfile(proc_dir + store_fname):
                    os.remove(proc_dir + store_fname)
                    shutil.copy2(store_file, proc_dir)
                else:
                    shutil.copy2(store_file, proc_dir)
            else:
                os.makedirs(proc_dir)
                shutil.copy2(store_file, proc_dir)
            os.chdir(proc_dir)
            if os.path.exists(proc_dir + dir_log) and os.path.isdir(proc_dir + dir_log):
                shutil.rmtree(proc_dir + dir_log)

            cf = tarfile.open(proc_dir + store_fname, 'r')
            cf.extractall(proc_dir)
            os.chdir(HOME)
            return proc_dir + dir_log

        except  Exception:
            os.chdir(HOME)
            exc_type, exc_value, exc_traceback = sys.exc_info()
            print "Se presento el error: %s %s %s0" % (exc_type, exc_value, exc_traceback)


class CreateReport:
    def __init__(self):
        pass

    @staticmethod
    def create_excel(db, servers, services_status, changes, incidents, downtime_logbook):

        date_report = months[int(datetime.datetime.now().strftime("%-m"))] + "_" + datetime.datetime.now().strftime(
            "%Y")
        excel = 'Reporte_disponibilidad_' + date_report + '.xlsx'
        excel_name = str('reports_manager/static/' + excel)
        workbook = xlsxwriter.Workbook(excel_name)
        worksheet = workbook.add_worksheet('Disponibilidad - ' + months[int(datetime.datetime.now().strftime("%-m"))])
        worksheet_d = workbook.add_worksheet(
            'Downtimes_table - ' + months[int(datetime.datetime.now().strftime("%-m"))])
        worksheet_d.write('A1', 'FECHA')
        worksheet_d.write('B1', 'TIEMPO DOWN TOTAL')
        worksheet_d.write('C1', 'BITACORAS CON DOWNTIME')
        worksheet_d.write('D1', 'SERVIDOR')
        tmp = 2

        format_header = workbook.add_format({
            'bold': 1,
            'border': 1,
            'border_color': '#8C8C8C',
            'align': 'Left',
            'font_color': 'white',
            'font_size': 9,
            'text_wrap': 1,
            'font_name': 'Constantia',
            'bg_color': '#4F81BD'})

        format_header2 = workbook.add_format({
            'bold': 1,
            'border': 1,
            'border_color': '#FFFFFF',
            'align': 'center',
            'font_color': 'white',
            'font_size': 9,
            'text_wrap': 1,
            'font_name': 'Constantia',
            'bg_color': '#4F81BD',
        })

        format_normal = workbook.add_format({
            'bold': 0,
            'align': 'Left',
            'valign': 'Left',
            'font_size': 9,
            'font_name': 'Constantia',
            'text_wrap': 1,
        })

        format_normal1 = workbook.add_format({
            'bold': 0,
            'align': 'Left',
            'valign': 'Left',
            'font_size': 9,
            'font_name': 'Constantia',
            'text_wrap': 1,
            'bg_color': '#DCE6F2',
            'border': 1,
            'border_color': '#FFFFFF',

        })

        format_inventario_content = workbook.add_format({
            'bold': False,
            'align': 'Center',
            'valign': 'vcenter',
            'font_size': 8,
            'font_name': 'Constantia',
            'text_wrap': 1,
            'bg_color': '#DCE6F2',
            'border': 1,
            'border_color': '#FFFFFF',
            'num_format': '0',
        })

        format_inventario_content1 = workbook.add_format({
            'bold': False,
            'align': 'Center',
            'valign': 'vcenter',
            'font_size': 8,
            'font_name': 'Constantia',
            'text_wrap': 1,
            'bg_color': '#DCE6F2',
            'border': 1,
            'border_color': '#FFFFFF',
            'num_format': '0',
        })

        format_inventario_content2 = workbook.add_format({
            'bold': False,
            'align': 'Center',
            'valign': 'vcenter',
            'font_size': 8,
            'font_name': 'Constantia',
            'text_wrap': 1,
            'bg_color': '#DCE6F2',
            'border': 1,
            'border_color': '#FFFFFF',
            'num_format': '0',
        })

        # 1 section
        cstatus = ""
        csl = ""
        worksheet.insert_image("A1", "reports_manager/static/logo.png", {'x_scale': 0.2, 'y_scale': 0.2})
        worksheet.set_column('A:A', 40)
        worksheet.merge_range('A5:C5', 'Soporte a licencias', format_normal)
        worksheet.merge_range('A6:C6', 'Reporte de resultado del servicio', format_normal)
        worksheet.write('A8', 'Mes', format_header)
        worksheet.write('B8', months[int(datetime.datetime.now().strftime("%-m"))], format_normal)
        worksheet.write('A9', 'Dias Mes', format_header)
        worksheet.write_number('B9', calendar.monthrange(int(datetime.datetime.now().strftime("%Y")),
                                                         int(datetime.datetime.now().strftime("%-m")))[1],
                               format_normal)
        worksheet.write('A10', 'Monto de \"Soporte de Licencias\" del Mes', format_header)
        worksheet.write('A11', 'Minutos Totales del Mes Por Servidor', format_header)
        worksheet.write_formula('B11', '=B9*1440', format_normal)

        worksheet.write('A14', 'CANT. SERV.', format_header2)
        worksheet.write_formula('A18', "=SUM(A15:A17)", format_header)
        worksheet.write('B14', 'Ambiente', format_header2)
        worksheet.write('B18', 'DTM', format_header2)
        worksheet.write('C18', '', format_header2)
        worksheet.write('D18', '', format_header2)
        worksheet.write('E18', '', format_header2)
        worksheet.write('B15', 'PRODUCCION', format_inventario_content)
        worksheet.write('B16', 'UAT', format_inventario_content)
        worksheet.write('B17', 'STRESS', format_inventario_content)
        worksheet.write('C14', 'SLA_Ambiente', format_header2)
        format_inventario_content1.set_num_format(0x0a)
        worksheet.write('C15', '99.99%', format_inventario_content1)
        worksheet.write('C16', '95.00%', format_inventario_content1)
        worksheet.write('C17', '95.00%', format_inventario_content1)

        worksheet.write('D14', 'SL_Ambiente', format_header2)
        format_inventario_content1.set_num_format(0x0a)
        worksheet.write_formula('D15', "=O3", format_inventario_content1)
        worksheet.write_formula('D16', "=AB3", format_inventario_content1)
        worksheet.write_formula('D17', "=AO3", format_inventario_content1)
        worksheet.write('E14', 'PFD', format_header2)
        format_inventario_content1.set_num_format(0x0a)
        worksheet.write('E15', '70%', format_inventario_content1)
        worksheet.write('E16', '20%', format_inventario_content1)
        worksheet.write('E17', '10%', format_inventario_content1)

        format_disclaimer = workbook.add_format({
            'bold': True,
            'align': 'Left',
            'valign': 'vcenter',
            'font_size': 8,
            'text_wrap': True,
            'font_name': 'Constantia',
        })
        message = u""""""
        worksheet.merge_range('A20:E21', message, format_disclaimer)

        worksheet.write('A23', 'NOMBRE', format_header2)
        worksheet.write('B23', 'ROL', format_header2)
        worksheet.merge_range('C23:D23', 'FIRMA', format_header2)

        # seccion production
        format_inventario = workbook.add_format({
            'bold': True,
            'align': 'Center',
            'valign': 'vcenter',
            'font_size': 8,
            'font_name': 'Constantia',
        })

        contract = u""""""
        month_total_day = calendar.monthrange(int(datetime.datetime.now().strftime("%Y")),
                                              int(datetime.datetime.now().strftime("%-m")))[1]
        month = months[int(datetime.datetime.now().strftime("%-m"))]
        year = datetime.datetime.now().strftime("%Y")
        period = u"Perí­odo del 1 de   " + str(month) + u"   al   " \
                 + str(month_total_day) + u"  de  " + str(month) + u"  del  " + str(year)
        worksheet.merge_range('G2:J2', 'Inventario Productivo', format_inventario)
        worksheet.merge_range('G5:R5', contract, format_inventario)
        worksheet.merge_range('G6:R6', period, format_inventario)
        worksheet.write("G7", "AMBIENTE", format_header2)
        worksheet.write("H7", "HOSTNAME", format_header2)
        worksheet.write("I7", "STATUS", format_header2)
        worksheet.write("J7", "IP", format_header2)
        worksheet.write("K7", "Plataforma", format_header2)
        worksheet.write("L7", "Grupo de Servicio", format_header2)
        worksheet.write("M7", "SERVICIO", format_header2)
        worksheet.write("N7", "Minutos de Tiempo Fuera", format_header2)
        worksheet.write("O7", "Minutos de Mantenimiento", format_header2)
        worksheet.write("P7", "SL_Componente", format_header2)
        worksheet.write("Q7", "Fechas Downtime", format_header2)
        worksheet.write("R7", "Archivos", format_header2)
        worksheet.write("S7", "Comentarios", format_header2)

        q = db.session.query(servers).filter(servers.environment == "produccion").order_by(servers.id.asc()).all()

        aux = 8

        number_month = datetime.datetime.now().strftime("%-m")
        number_days = calendar.monthrange(int(datetime.datetime.now().strftime("%Y")),
                                          int(number_month))[1]
        if int(number_month) < 10:
            month_ini = datetime.datetime.now().strftime("01-" + "0" + number_month + "-%Y")
            month_end = datetime.datetime.now().strftime(str(number_days) + "0" + number_month + "-" + "%Y-")
        else:
            month_ini = datetime.datetime.now().strftime("01-" + number_month + "-" + str(year))
            month_end = datetime.datetime.now().strftime(str(number_days) + "-" + number_month + "-" + str(year))

        aux_mini = datetime.datetime.strptime(month_ini, "%d-%m-%Y")
        aux_mend = datetime.datetime.strptime(month_end, "%d-%m-%Y")
        for row in q:
            down_total = []
            d_date = []
            time_c = []
            time_i = []
            comments_pc = []
            comments_pi = []
            fdp_tmp = []
            qcp_tmp = []
            qip_tmp = []
            qddp_tmp = []
            tmp_d = []

            q1 = db.session.query(services_status).filter(
                and_(services_status.servers_id == int(row.id),
                     between(services_status.date, str(month_ini), str(month_end)))).all()

            for row1 in q1:
                rowa = "A" + str(tmp)
                rowb = "B" + str(tmp)
                rowc = "C" + str(tmp)
                rowd = "D" + str(tmp)
                aux_date = datetime.datetime.strptime(row1.date, "%d-%m-%Y")
                if aux_mini <= aux_date <= aux_mend:
                    worksheet_d.write(rowa, row1.date)
                    worksheet_d.write(rowb, row1.time_down)
                    worksheet_d.write(rowd, row1.hostname)

                if row1.files_down != '':
                    if aux_mini <= aux_date <= aux_mend:
                        down_total.append(int(row1.time_down))
                        d_date.append(str(row1.date))
                        fd_dsp = ast.literal_eval(row1.files_down)
                        fdp_tmp.append(bubble_sort.sort(fd_dsp.keys()))
                        worksheet_d.write(rowc, str(fd_dsp))
                else:
                    pass
                    # worksheet_d.write(rowc, "")
                tmp += 1
            d_date = bubble_sort.sort(d_date)
            for dc in d_date:
                qc = db.session.query(changes).filter(changes.date_ini == dc.replace("-", "/")).all()
                qcp_tmp.append(qc)
                qi = db.session.query(incidents).filter(incidents.datecab == dc).all()
                qip_tmp.append(qi)
                qddp = db.session.query(downtime_logbook).filter(downtime_logbook.date_downtime == dc).all()
                tmp_t = []

                for k in qddp:
                    if re.search(row.hostname, k.file_name):
                        tmp_t.append(k.time_downtime)
                g = str(sum(tmp_t)) + "  min " + " el " + str(dc.split("-")[0]) + "-" + str(
                    months[int(dc.split("-")[1])])
                qddp_tmp.append(g)
                tmp_d.append(str(dc.split("-")[0]) + "-" + str(months[int(dc.split("-")[1])]))

                for qc_aux in qcp_tmp:
                    for c in qc_aux:
                        for server in ast.literal_eval(c.listing):
                            if re.match(row.hostname, server):
                                #             t1 = datetime.datetime.strptime(c.time_ini, "%H:%M")
                                #             t2 = datetime.datetime.strptime(c.time_end, "%H:%M")
                                #             time_aux = t2 - t1
                                #             time_c.append((time_aux.seconds / 60))
                                #             comments_pc.append(c.notes)
                                descr = str(dc.split("-")[0]) + "-" + str(
                                    months[int(dc.split("-")[1])]) + "/Ejecución de RFC " + str(
                                    c.reference) + "/" + str(c.description)
                                comments_pc.append(descr)
                                #
                                # for qi_aux in qip_tmp:
                                #     for i in qi_aux:
                                #         if row.hostname in i.listing:
                                #             t1 = datetime.datetime.strptime(i.time_ini, "%H:%M")
                                #             t2 = datetime.datetime.strptime(i.time_end, "%H:%M")
                                #             time_aux = t2 - t1
                                #             time_i.append((time_aux.seconds / 60))
                                #            comments_pi.append(i.notes)

            # ttcip = sum(time_c) + sum(time_i)
            ttcip = sum(down_total)
            dtp = sum(down_total)
            fdp = '\n'.join(map(str, fdp_tmp)).replace('[', '').replace(']', '').replace('\'', '').replace('\n',
                                                                                                           '').replace(
                ',', '\n')
            ddp = '\n'.join(map(str, qddp_tmp))
            commentsp = '\n'.join(map(str, comments_pc)) + "\n" + '\n'.join(map(str, comments_pi)) + "\n"
            chostname = "H" + str(aux)
            cenvironment = "G" + str(aux)
            cstatus = "I" + str(aux)
            cip = "J" + str(aux)
            cplatform = "K" + str(aux)
            gservices = "L" + str(aux)
            cservices = "M" + str(aux)
            ctime_down = "N" + str(aux)
            cmaintenance = "O" + str(aux)
            csl = "P" + str(aux)
            cdates_down = "Q" + str(aux)
            cfiles_down = "R" + str(aux)
            ccoments = "S" + str(aux)
            csl_form = '=($O$1-(' + ctime_down + '-' + cmaintenance + '))/$O$1'
            worksheet.write(cenvironment, u"PRODUCCION", format_inventario_content)
            worksheet.write(chostname, row.hostname, format_inventario_content)
            worksheet.write(cstatus, str(row.status.encode('utf8')).upper(), format_inventario_content)
            worksheet.write(cip, row.ip_prod, format_inventario_content)
            worksheet.write(cplatform, row.platform, format_inventario_content)
            worksheet.write(gservices, row.service_group, format_inventario_content)
            worksheet.write(cservices, row.service, format_inventario_content)
            format_inventario_content.set_num_format('#,##0.000')
            worksheet.write(ctime_down, dtp, format_inventario_content)
            worksheet.write(cfiles_down, str(fdp), format_inventario_content)
            if len(str(fdp)) != 0:
                worksheet.write(cdates_down, str(ddp), format_inventario_content)
            else:
                worksheet.write(cdates_down, "", format_inventario_content)

            worksheet.write(cmaintenance, str(ttcip), format_inventario_content)
            worksheet.write(ccoments, str(commentsp), format_inventario_content)
            worksheet.write_formula(csl, csl_form, format_inventario_content)
            if len(commentsp.strip()) != 0:
                worksheet.write(cmaintenance, str(ttcip), format_inventario_content)
                worksheet.write(ccoments, str(commentsp), format_inventario_content)
            else:
                worksheet.write(cmaintenance, 0, format_inventario_content)
                comments2 = '\n'.join(map(str, tmp_d))
                worksheet.write(ccoments, str(comments2), format_inventario_content)
            aux += 1

        copera = "H" + str(aux)
        worksheet.write(copera, "TOTAL", format_inventario)
        ctopera = "I" + str(aux)
        fopera = '=COUNTIF(I8:' + cstatus.strip() + ', \"OPERATIVO\")'
        worksheet.write_formula(ctopera, fopera, format_inventario_content)
        worksheet.write_formula("A15", "=" + ctopera, format_normal1)
        # calta = "H" + str(aux + 1)
        # worksheet.write(calta, "ALTA", format_inventario)
        # ctalta = "I" + str(aux + 1)
        # ftalta = '=COUNTIF(I8:' + cstatus.strip() + ', \"ALTA\")'
        # worksheet.write_formula(ctalta, ftalta, format_inventario_content)
        # ctotal = "H" + str(aux + 2)
        # cttotal = "I" + str(aux + 2)
        ctime_downtotal = "N" + str(aux)
        ctime_just = "O" + str(aux)
        #        worksheet.write(ctotal, "TOTAL", format_inventario)
        #        worksheet.write_formula(cttotal, 'SUM(' + ctopera + ":" + ctalta + ')', format_inventario_content)
        #       worksheet.write_formula("A15", "=" + cttotal, format_inventario_content)
        worksheet.write_formula(ctime_downtotal, '=SUM(' + 'M8' + ':' + 'M' + str(aux - 1) + ')',
                                format_inventario_content)
        worksheet.write_formula(ctime_just, '=SUM(' + 'N8' + ':' + 'N' + str(aux - 1) + ')', format_inventario_content)
        worksheet.write("N1", "t:", format_header)
        worksheet.write_formula("O1", '=B11', format_normal1)
        worksheet.write("N2", "SLA:", format_header)
        worksheet.write("O2", "99.90%", format_inventario_content1)
        worksheet.write("N3", "Promedio:", format_header)
        format_inventario_content1.set_num_format(0x0a)
        worksheet.write_formula("O3", '=AVERAGE(P8:' + csl + ')', format_inventario_content1)

        # seccion 3 

        contract = u""""""
        month_total_day = calendar.monthrange(int(datetime.datetime.now().strftime("%Y")),
                                              int(datetime.datetime.now().strftime("%-m")))[1]
        month = months[int(datetime.datetime.now().strftime("%-m"))]
        year = datetime.datetime.now().strftime("%Y")
        period = u"Perí­odo del 1 de   " + str(month) + u"   al   " \
                 + str(month_total_day) + u"  de  " + str(month) + u"  del  " + str(year)
        worksheet.merge_range('T2:W2', 'Inventario UAT', format_inventario)
        worksheet.merge_range('T5:AE5', contract, format_inventario)
        worksheet.merge_range('T6:AE6', period, format_inventario)
        worksheet.write("U7", "AMBIENTE", format_header2)
        worksheet.write("V7", "HOSTNAME", format_header2)
        worksheet.write("W7", "STATUS", format_header2)
        worksheet.write("X7", "IP", format_header2)
        worksheet.write("Y7", "Plataforma", format_header2)
        worksheet.write("Z7", "Grupo de servicio")
        worksheet.write("AA7", "SERVICIO", format_header2)
        worksheet.write("AB7", "Minutos de Tiempo Fuera", format_header2)
        worksheet.write("AC7", "Minutos de Mantenimiento", format_header2)
        worksheet.write("AD7", "SL_Componente", format_header2)
        worksheet.write("AE7", "Fechas Downtime", format_header2)
        worksheet.write("AF7", "Archivos", format_header2)
        worksheet.write("AG7", "Comentarios", format_header2)

        q = db.session.query(servers).filter(servers.environment == "uat").order_by(servers.id.asc()).all()
        aux = 8

        for row in q:
            down_total = []
            d_date = []
            time_c = []
            time_i = []
            comments_uc = []
            comments_ui = []
            fdu_tmp = []
            qcu_tmp = []
            qiu_tmp = []
            qddu_tmp = []
            tmp_d = []
            q1 = db.session.query(services_status).filter(
                and_(services_status.servers_id == int(row.id),
                     between(services_status.date, str(month_ini), str(month_end)))).all()

            for row1 in q1:
                rowa = "A" + str(tmp)
                rowb = "B" + str(tmp)
                rowc = "C" + str(tmp)
                rowd = "D" + str(tmp)
                aux_date = datetime.datetime.strptime(row1.date, "%d-%m-%Y")
                if aux_mini <= aux_date <= aux_mend:
                    worksheet_d.write(rowa, row1.date)
                    worksheet_d.write(rowb, row1.time_down)
                    worksheet_d.write(rowd, row1.hostname)

                if len(row1.files_down) != 0:
                    if aux_mini <= aux_date <= aux_mend:
                        down_total.append(int(row1.time_down))
                        d_date.append(str(row1.date))
                        fd_dsu = ast.literal_eval(row1.files_down)
                        fdu_tmp.append(bubble_sort.sort(fd_dsu.keys()))
                        worksheet_d.write(rowc, str(fd_dsu))
                else:
                    worksheet_d.write(rowc, "")
                tmp += 1
            d_date = bubble_sort.sort(d_date)
            for dc in d_date:
                qc = db.session.query(changes).filter(changes.date_ini == dc.replace("-", "/")).all()
                qcu_tmp.append(qc)
                qi = db.session.query(incidents).filter(incidents.datecab == dc).all()
                qiu_tmp.append(qi)
                qddu = db.session.query(downtime_logbook).filter(downtime_logbook.date_downtime == dc).all()
                tmp_t = []
                for k in qddu:
                    if re.search(row.hostname, k.file_name):
                        tmp_t.append(k.time_downtime)
                g = str(sum(tmp_t)) + "  min " + " el " + str(dc.split("-")[0]) + "-" + str(
                    months[int(dc.split("-")[1])])
                qddu_tmp.append(g)
                tmp_d.append(str(dc.split("-")[0]) + "-" + str(months[int(dc.split("-")[1])]))

                for qc_aux in qcu_tmp:
                    for c in qc_aux:
                        for server in ast.literal_eval(c.listing):
                            if re.match(row.hostname, server):
                                #             t1 = datetime.datetime.strptime(c.time_ini, "%H:%M")
                                #             t2 = datetime.datetime.strptime(c.time_end, "%H:%M")
                                #             time_aux = t2 - t1
                                #             time_c.append((time_aux.seconds / 60))ç
                                descru = str(dc.split("-")[0]) + "-" + str(months[int(dc.split("-")[1])]) + \
                                         "/Ejecución de RFC  " + str(c.reference) + "/" + str(c.description)
                                comments_uc.append(descru)
                                #             comments_uc.append(c.notes)
                                #
                                # for qi_aux in qiu_tmp:
                                #     for i in qi_aux:
                                #         if row.hostname in i.listing:
                                #             t1 = datetime.datetime.strptime(i.time_ini, "%H:%M")
                                #             t2 = datetime.datetime.strptime(i.time_end, "%H:%M")
                                #             time_aux = t2 - t1
                                #             time_i.append((time_aux.seconds / 60))
                                #             comments_ui.append(i.notes)

            # ttciu = sum(time_c) + sum(time_i)
            ttciu = sum(down_total)
            dtu = sum(down_total)
            fdu = '\n'.join(map(str, fdu_tmp)).replace('[', '').replace(']', '').replace('\'', '')
            ddu = '\n'.join(map(str, qddu_tmp))
            commentsu = '\n'.join(map(str, comments_uc)) + "\n" + '\n'.join(map(str, comments_ui)) + "\n"
            chostname = "V" + str(aux)
            cenvironment = "U" + str(aux)
            cstatus = "W" + str(aux)
            cip = "X" + str(aux)
            cplatform = "Y" + str(aux)
            gservices = "Z" + str(aux)
            cservices = "AA" + str(aux)
            ctime_down = "AB" + str(aux)
            cmaintenance = "AC" + str(aux)
            csl = "AD" + str(aux)
            cdates_down = "AE" + str(aux)
            cfiles_down = "AF" + str(aux)
            ccoments = "AG" + str(aux)
            csl_form = '=($O$1-(' + ctime_down + '-' + cmaintenance + '))/$O$1'
            worksheet.write(cenvironment, "UAT", format_inventario_content)
            worksheet.write(chostname, row.hostname, format_inventario_content)
            worksheet.write(cstatus, str(row.status).upper(), format_inventario_content)
            worksheet.write(cip, row.ip_prod, format_inventario_content)
            worksheet.write(cplatform, row.platform, format_inventario_content)
            worksheet.write(gservices, row.service_group, format_inventario_content)
            worksheet.write(cservices, row.service, format_inventario_content)
            format_inventario_content.set_num_format('#,##0.000')
            worksheet.write_formula(csl, csl_form, format_inventario_content)
            worksheet.write(ctime_down, dtu, format_inventario_content)
            worksheet.write(cfiles_down, str(fdu), format_inventario_content)

            if len(str(fdu)) != 0 or fdu != "":
                worksheet.write(cdates_down, str(ddu), format_inventario_content)
            else:
                worksheet.write(cdates_down, "", format_inventario_content)

            if len(str(commentsu).strip()) != 0:
                worksheet.write(cmaintenance, str(ttciu), format_inventario_content)
                worksheet.write(ccoments, str(commentsu), format_inventario_content)
            else:
                worksheet.write(cmaintenance, 0, format_inventario_content)
                comments2 = '\n'.join(map(str, tmp_d))
                worksheet.write(ccoments, str(comments2), format_inventario_content)

            aux += 1

        copera = "V" + str(aux)
        worksheet.write(copera, "TOTAL", format_inventario)
        ctopera = "W" + str(aux)
        fopera = '=COUNTIF(V8:' + cstatus.strip() + ', \"OPERATIVO\")'
        worksheet.write_formula(ctopera, fopera, format_inventario_content)
        worksheet.write_formula('A16', "=" + ctopera, format_normal1)
        #        calta = "U" + str(aux + 1)
        #       worksheet.write(calta, "ALTA", format_inventario)
        # ctalta = "V" + str(aux + 1)
        # ftalta = '=COUNTIF(V8:' + cstatus.strip() + ', \"ALTA\")'
        # worksheet.write_formula(ctalta, ftalta, format_inventario_content)
        #        ctotal = "U" + str(aux + 2)
        #        cttotal = "V" + str(aux + 2)
        ctime_downtotal = "AB" + str(aux)
        ctime_just = "AC" + str(aux)
        # worksheet.write(ctotal, "TOTAL", format_inventario)
        # worksheet.write_formula(cttotal, 'SUM(' + ctopera + ":" + ctalta + ')', format_inventario_content)
        # worksheet.write_formula("A16", "=" + cttotal, format_inventario_content)
        worksheet.write_formula(ctime_downtotal, "=SUM(AB8" + ":" + ("AB" + str(aux - 1)) + ")",
                                format_inventario_content)
        worksheet.write_formula(ctime_just, "=SUM(AC8" + ":" + ("AC" + str(aux - 1)) + ")", format_inventario_content)
        worksheet.write("AA1", "t:", format_header)
        worksheet.write_formula("AB1", '=B11', format_normal1)
        worksheet.write("AA2", "SLA:", format_header)
        worksheet.write("AB2", "95.00%", format_inventario_content1)
        worksheet.write("AA3", "Promedio:", format_header)
        format_inventario_content1.set_num_format(0x0a)
        worksheet.write_formula("AB3", '=AVERAGE(AD8:' + csl + ")", format_inventario_content1)

        # seccion 4 

        contract = u""""""
        month_total_day = calendar.monthrange(int(datetime.datetime.now().strftime("%Y")),
                                              int(datetime.datetime.now().strftime("%-m")))[1]
        month = months[int(datetime.datetime.now().strftime("%-m"))]
        year = datetime.datetime.now().strftime("%Y")
        period = u"Perí­odo del 1 de   " + str(month) + u"   al   " \
                 + str(month_total_day) + u"  de  " + str(month) + u"  del  " + str(year)
        worksheet.merge_range('AG2:AJ2', 'Inventario Stress', format_inventario)
        worksheet.merge_range('AG5:AR5', contract, format_inventario)
        worksheet.merge_range('AG6:AR6', period, format_inventario)
        worksheet.write("AJ7", "AMBIENTE", format_header2)
        worksheet.write("AK7", "HOSTNAME", format_header2)
        worksheet.write("AL7", "STATUS", format_header2)
        worksheet.write("AM7", "IP", format_header2)
        worksheet.write("AN7", "Plataforma", format_header2)
        worksheet.write("AO7", "Grupo de servicio", format_header2)
        worksheet.write("AP7", "SERVICIO", format_header2)
        worksheet.write("AQ7", "Minutos de Tiempo Fuera", format_header2)
        worksheet.write("AR7", "Minutos de Mantenimiento", format_header2)
        worksheet.write("AS7", "SL_Componente", format_header2)
        worksheet.write("AT7", "Fechas Downtime", format_header2)
        worksheet.write("AU7", "Archivos", format_header2)
        worksheet.write("AV7", "Comentarios", format_header2)

        q = db.session.query(servers).filter(servers.environment == "stress").order_by(servers.id.asc()).all()
        aux = 8

        for row in q:
            down_total = []
            d_date = []
            # time_c = []
            # time_i = []
            comments_sc = []
            comments_si = []
            fds_tmp = []
            qcs_tmp = []
            qis_tmp = []
            qddd_tmp = []
            tmp_d = []
            q1 = db.session.query(services_status).filter(
                and_(services_status.servers_id == int(row.id),
                     between(services_status.date, str(month_ini), str(month_end)))).all()
            for row1 in q1:
                rowa = "A" + str(tmp)
                rowb = "B" + str(tmp)
                rowc = "C" + str(tmp)
                rowd = "D" + str(tmp)
                aux_date = datetime.datetime.strptime(row1.date, "%d-%m-%Y")
                if aux_mini <= aux_date <= aux_mend:
                    worksheet_d.write(rowb, row1.time_down)
                    worksheet_d.write(rowd, row1.hostname)

                if len(row1.files_down) != 0:
                    if aux_mini <= aux_date <= aux_mend:
                        down_total.append(int(row1.time_down))
                        d_date.append(str(row1.date))
                        worksheet_d.write(rowa, row1.date)
                        fd_dss = ast.literal_eval(row1.files_down)
                        fds_tmp.append(bubble_sort.sort(fd_dss.keys()))
                        worksheet_d.write(rowc, str(fd_dss))
                else:
                    worksheet_d.write(rowa, "")
                    worksheet_d.write(rowc, "")
                tmp += 1
            d_date = bubble_sort.sort(d_date)
            for dc in d_date:
                qc = db.session.query(changes).filter(changes.date_ini == dc.replace("-", "/")).all()
                qcs_tmp.append(qc)
                qi = db.session.query(incidents).filter(incidents.datecab == dc).all()
                qis_tmp.append(qi)
                qdds = db.session.query(downtime_logbook).filter(downtime_logbook.date_downtime == dc).all()
                tmp_t = []
                for k in qdds:
                    if re.search(row.hostname, k.file_name):
                        tmp_t.append(k.time_downtime)
                g = str(sum(tmp_t)) + "  min " + " el " + str(dc.split("-")[0]) + "-" + str(
                    months[int(dc.split("-")[1])])
                qddd_tmp.append(g)
                tmp_d.append(str(dc.split("-")[0]) + "-" + str(months[int(dc.split("-")[1])]))

            for qc_aux in qcs_tmp:
                for c in qc_aux:
                    t1 = datetime.datetime.strptime(c.time_ini, "%H:%M")
                    t2 = datetime.datetime.strptime(c.time_end, "%H:%M")
                    for server in ast.literal_eval(c.listing):
                        if re.match(row.hostname, server) and re.match(str(t1), server):
                            #             t1 = datetime.datetime.strptime(c.time_ini, "%H:%M")
                            #             t2 = datetime.datetime.strptime(c.time_end, "%H:%M")
                            #             time_aux = t2 - t1
                            #             time_c.append((time_aux.seconds / 60))
                            descrs = str(dc.split("-")[0]) + "-" + str(months[int(dc.split("-")[1])]) + \
                                     "/Ejecución de RFC " + str(c.reference) + "/" + str(c.description)
                            comments_sc.append(descrs)
            # comments_sc.append(c.notes)

            # for qi_aux in qis_tmp:
            #     for i in qi_aux:
            #
            #         if row.hostname in i.listing:
            #             t1 = datetime.datetime.strptime(i.time_ini, "%H:%M")
            #             t2 = datetime.datetime.strptime(i.time_end, "%H:%M")
            #             time_aux = t2 - t1
            #             time_i.append((time_aux.seconds / 60))
            #             comments_si.append(i.notes)

            # ttcis = sum(time_c) + sum(time_i)
            ttcis = sum(down_total)
            dts = sum(down_total)
            fds = '\n'.join(map(str, fds_tmp)).replace('[', '').replace(']', '').replace('\'', '')
            dds = '\n'.join(map(str, qddd_tmp))
            commentss = '\n'.join(map(str, comments_sc)) + "\n" + '\n'.join(map(str, comments_si)) + "\n"
            chostname = "AK" + str(aux)
            cenvironment = "AJ" + str(aux)
            cstatus = "AL" + str(aux)
            cip = "AM" + str(aux)
            cplatform = "AN" + str(aux)
            gservices = "AM" + str(aux)
            cservices = "AO" + str(aux)
            ctime_down = "AP" + str(aux)
            cmaintenance = "AR" + str(aux)
            csl = "AS" + str(aux)
            cdates_down = "AT" + str(aux)
            cfiles_down = "AU" + str(aux)
            ccoments = "AV" + str(aux)
            csl_form = '=($O$1-(' + ctime_down + '-' + cmaintenance + '))/$O$1'

            worksheet.write(cenvironment, "STRESS", format_inventario_content)
            worksheet.write(chostname, row.hostname, format_inventario_content)
            worksheet.write(cstatus, str(row.status).upper(), format_inventario_content)
            worksheet.write(cip, row.ip_prod, format_inventario_content)
            worksheet.write(cplatform, row.platform, format_inventario_content)
            worksheet.write(gservices, row.service_group, format_inventario_content)
            worksheet.write(cservices, row.service, format_inventario_content)
            format_inventario_content.set_num_format('#,##0.000')
            worksheet.write_formula(csl, csl_form, format_inventario_content)
            worksheet.write(ctime_down, dts, format_inventario_content)
            worksheet.write(cfiles_down, str(fds), format_inventario_content)
            if len(str(fds)) != 0:

                worksheet.write(cdates_down, str(dds), format_inventario_content)
            else:

                worksheet.write(cdates_down, "", format_inventario_content)
            if len(commentss.strip()) != 0:
                worksheet.write(cmaintenance, str(ttcis), format_inventario_content)
                worksheet.write(ccoments, str(commentss), format_inventario_content)
            else:
                worksheet.write(cmaintenance, 0, format_inventario_content)
                comments2 = '\n'.join(map(str, tmp_d))
                worksheet.write(ccoments, str(comments2), format_inventario_content)

            aux += 1

        # copera = "AH" + str(aux)
        #        worksheet.write(copera, "OPERACIONAL", format_inventario)
        #        calta = "AH" + str(aux + 1)
        #        worksheet.write(calta, "ALTA", format_inventario)
        ctotal = "AJ" + str(aux + 2)
        worksheet.write(ctotal, "TOTAL", format_inventario)

        #        copera = "AJ" + str(aux)
        #        worksheet.write(copera, "OPERACIONAL", format_inventario)
        ctopera = "AK" + str(aux)
        fopera = '=COUNTIF(AL8:' + cstatus.strip() + ', \"OPERATIVO\")'
        worksheet.write_formula(ctopera, fopera, format_inventario_content)
        calta = "AH" + str(aux + 1)
        worksheet.write(calta, "TOTAL", format_inventario)
        ctopera = "AI" + str(aux)
        fopera = '=COUNTIF(AI8:' + cstatus.strip() + ', \"OPERATIVO\")'
        worksheet.write_formula(ctopera, fopera, format_inventario_content)
        worksheet.write_formula('A17', "=" + ctopera, format_normal1)
        #        ctalta = "AI" + str(aux + 1)
        #       ftalta = '=COUNTIF(AI8:' + cstatus.strip() + ', \"ALTA\")'
        #       worksheet.write_formula(ctalta, ftalta, format_inventario_content)
        #        ctotal = "AH" + str(aux + 2)
        #        cttotal = "AI" + str(aux + 2)
        # worksheet.write(ctotal, "TOTAL", format_inventario)
        # worksheet.write_formula(cttotal, 'SUM(' + ctopera + ":" + ctalta + ')', format_inventario_content)
        # worksheet.write_formula("A17", "=" + cttotal, format_inventario_content)
        ctime_downtotal = "AQ" + str(aux)
        ctime_just = "AR" + str(aux)
        worksheet.write_formula(ctime_downtotal, "=SUM(AM8" + ":" + "AM" + str(aux - 1) + ")",
                                format_inventario_content)
        worksheet.write_formula(ctime_just, "=SUM(AN8" + ":" + "AN" + str(aux - 1) + ")", format_inventario_content)
        worksheet.write("AN1", "t:", format_header)
        worksheet.write_formula("AO1", '=B11', format_normal1)
        worksheet.write("AN2", "SLA:", format_header)
        worksheet.write("AO2", "95.00%", format_inventario_content1)
        worksheet.write("AN3", "Promedio:", format_header)
        format_inventario_content1.set_num_format(0x0a)
        worksheet.write_formula("AO3", '=AVERAGE(AS8:' + csl + ")", format_inventario_content1)

        workbook.close()
        return excel, excel_name

    @staticmethod
    def create_excel2(db, servers, services_status, changes, incidents, year, month1, downtime_logbook):

        date_report = months[int(month1)] + "_" + year
        excel = 'Reporte_disponibilidad_' + date_report + '.xlsx'
        excel_name = str('./reports_manager/static/' + excel)
        workbook = xlsxwriter.Workbook(excel_name)
        worksheet = workbook.add_worksheet('Disponibilidad - ' + months[int(month1)])
        worksheet_d = workbook.add_worksheet(
            'Downtimes_table - ' + months[int(month1)])
        worksheet_d.write('A1', 'FECHA')
        worksheet_d.write('B1', 'TIEMPO DOWN TOTAL')
        worksheet_d.write('C1', 'BITACORAS CON DOWNTIME')
        worksheet_d.write('D1', 'SERVIDOR')
        tmp = 2

        format_header = workbook.add_format({
            'bold': 1,
            'border': 1,
            'border_color': '#FFFFFF',
            'align': 'Left',
            'font_color': 'white',
            'font_size': 9,
            'text_wrap': 1,
            'font_name': 'Constantia',
            'bg_color': '#4F81BD'})

        format_header2 = workbook.add_format({
            'bold': 1,
            'border': 1,
            'border_color': '#FFFFFF',
            'align': 'center',
            'font_color': 'white',
            'font_size': 9,
            'text_wrap': 1,
            'font_name': 'Constantia',
            'bg_color': '#4F81BD',
        })

        format_normal = workbook.add_format({
            'bold': 0,
            'align': 'Left',
            'valign': 'Left',
            'font_size': 9,
            'font_name': 'Constantia',
            'text_wrap': 1,

        })

        format_normal1 = workbook.add_format({
            'bold': 0,
            'align': 'Left',
            'valign': 'Left',
            'font_size': 9,
            'font_name': 'Constantia',
            'text_wrap': 1,
            'bg_color': '#DCE6F2',
            'border': 1,
            'border_color': '#FFFFFF',

        })

        format_inventario_content = workbook.add_format({
            'bold': False,
            'align': 'Center',
            'valign': 'vcenter',
            'font_size': 8,
            'font_name': 'Constantia',
            'text_wrap': 1,
            'num_format': '0',
            'bg_color': '#DCE6F2',
            'border': 1,
            'border_color': '#FFFFFF',
        })

        format_inventario_content1 = workbook.add_format({
            'bold': False,
            'align': 'Center',
            'valign': 'vcenter',
            'font_size': 8,
            'font_name': 'Constantia',
            'text_wrap': 1,
            'bg_color': '#DCE6F2',
            'border': 1,
            'border_color': '#FFFFFF',
            'num_format': '0',
        })

        format_inventario_content2 = workbook.add_format({
            'bold': False,
            'align': 'Center',
            'valign': 'vcenter',
            'font_size': 8,
            'font_name': 'Constantia',
            'text_wrap': 1,
            'bg_color': '#DCE6F2',
            'border': 1,
            'border_color': '#FFFFFF',
            'num_format': '0',
        })

        # 1 section
        cstatus = ""
        csl = ""
        print os.getcwd()
        worksheet.insert_image("A1", "./reports_manager/static/.png", {'x_scale': 0.2, 'y_scale': 0.2})
        worksheet.set_column('A:A', 40)
        worksheet.merge_range('A5:C5', 'Soporte a licencias', format_normal)
        worksheet.merge_range('A6:C6', 'Reporte de resultado del servicio', format_normal)
        worksheet.write('A8', 'Mes', format_header)
        worksheet.write('B8', months[int(month1)], format_normal1)
        worksheet.write('A9', 'Dias Mes', format_header)
        worksheet.write_number('B9', calendar.monthrange(int(year), int(month1))[1], format_normal1)
        worksheet.write('A10', 'Monto de \"Soporte de Licencias\" del Mes', format_header)
        worksheet.write('B10', '', format_normal1)
        worksheet.write('A11', 'Minutos Totales del Mes Por Servidor', format_header)
        worksheet.write_formula('B11', '=B9*1440', format_normal1)

        worksheet.write('A14', 'CANT. SERV.', format_header2)
        worksheet.write_formula('A18', "=SUM(A15:A17)", format_header)
        worksheet.write('B14', 'Ambiente', format_header2)
        worksheet.write('B18', 'DTM', format_header2)
        worksheet.write('C18', '', format_header2)
        worksheet.write('D18', '', format_header2)
        worksheet.write('E18', '', format_header2)
        worksheet.write('B15', 'PRODUCCION', format_inventario_content)
        worksheet.write('B16', 'UAT', format_inventario_content)
        worksheet.write('B17', 'STRESS', format_inventario_content)
        worksheet.write('C14', 'SLA_Ambiente', format_header2)
        format_inventario_content1.set_num_format(0x0a)
        worksheet.write('C15', '99.99%', format_inventario_content1)
        worksheet.write('C16', '95.00%', format_inventario_content1)
        worksheet.write('C17', '95.00%', format_inventario_content1)

        worksheet.write('D14', 'SL_Ambiente', format_header2)
        format_inventario_content1.set_num_format(0x0a)
        worksheet.write_formula('D15', "=O3", format_inventario_content1)
        worksheet.write_formula('D16', "=AB3", format_inventario_content1)
        worksheet.write_formula('D17', "=AO3", format_inventario_content1)
        worksheet.write('E14', 'PFD', format_header2)
        format_inventario_content1.set_num_format(0x0a)
        worksheet.write('E15', '70%', format_inventario_content1)
        worksheet.write('E16', '20%', format_inventario_content1)
        worksheet.write('E17', '10%', format_inventario_content1)

        format_disclaimer = workbook.add_format({
            'bold': True,
            'align': 'Left',
            'valign': 'vcenter',
            'font_size': 8,
            'text_wrap': True,
            'font_name': 'Constantia',
        })
        message = u"""."""
        worksheet.merge_range('A20:E21', message, format_disclaimer)

        worksheet.write('A23', 'NOMBRE', format_header2)
        worksheet.write('B23', 'ROL', format_header2)
        worksheet.merge_range('C23:D23', 'FIRMA', format_header2)

        # seccion production
        format_inventario = workbook.add_format({
            'bold': True,
            'align': 'Center',
            'valign': 'vcenter',
            'font_size': 8,
            'font_name': 'Constantia',
        })

        contract = u""""""
        month_total_day = calendar.monthrange(int(year), int(month1))[1]
        month = months[int(month1)]
        period = u"Perí­odo del 1 de   " + str(month) + u"   al   " \
                 + str(month_total_day) + u"  de  " + str(month) + u"  del  " + str(year)
        worksheet.merge_range('G2:J2', 'Inventario Productivo', format_inventario)
        worksheet.merge_range('G5:R5', contract, format_inventario)
        worksheet.merge_range('G6:R6', period, format_inventario)
        worksheet.write("G7", "AMBIENTE", format_header2)
        worksheet.write("H7", "HOSTNAME", format_header2)
        worksheet.write("I7", "STATUS", format_header2)
        worksheet.write("J7", "IP", format_header2)
        worksheet.write("K7", "Plataforma", format_header2)
        worksheet.write("L7", "Grupo de Servicio", format_header2)
        worksheet.write("M7", "SERVICIO", format_header2)
        worksheet.write("N7", "Minutos de Tiempo Fuera", format_header2)
        worksheet.write("O7", "Minutos de Mantenimiento", format_header2)
        worksheet.write("P7", "SL_Componente", format_header2)
        worksheet.write("Q7", "Fechas Downtime", format_header2)
        worksheet.write("R7", "Archivos", format_header2)
        worksheet.write("S7", "Comentarios", format_header2)

        q = db.session.query(servers).filter(servers.environment == "produccion").order_by(servers.id.asc()).all()

        aux = 8
        number_month = month1

        number_days = calendar.monthrange(int(year), int(number_month))[1]
        #
        # if int(number_month) < 10:
        #     month_ini = datetime.datetime.now().strftime("01-" + "0" + number_month + "-" + str(year))
        #     month_end = datetime.datetime.now().strftime(str(number_days) + "-0" + number_month + "-" + str(year))
        # else:
        month_ini = datetime.datetime.now().strftime("01-" + number_month + "-" + str(year))
        month_end = datetime.datetime.now().strftime(str(number_days) + "-" + number_month + "-" + str(year))

        aux_mini = datetime.datetime.strptime(month_ini, "%d-%m-%Y")
        aux_mend = datetime.datetime.strptime(month_end, "%d-%m-%Y")
        for row in q:
            down_total = []
            d_date = []
            comments_pc = []
            comments_pi = []
            fdp_tmp = []
            qcp_tmp = []
            qip_tmp = []
            qddp_tmp = []
            tmp_d = []

            q1 = db.session.query(services_status).filter(
                and_(services_status.hostname == row.hostname, between(services_status.date,
                                                                       str(month_ini), str(month_end)))).all()

            for row1 in q1:
                rowa = "A" + str(tmp)
                rowb = "B" + str(tmp)
                rowc = "C" + str(tmp)
                rowd = "D" + str(tmp)

                aux_date = datetime.datetime.strptime(row1.date, "%d-%m-%Y")

                if aux_mini <= aux_date <= aux_mend:
                    worksheet_d.write(rowa, row1.date)
                    worksheet_d.write(rowb, row1.time_down)
                    worksheet_d.write(rowd, row1.hostname)

                if row1.files_down != '':
                    if aux_mini <= aux_date <= aux_mend:
                        down_total.append(int(row1.time_down))
                        d_date.append(str(row1.date))
                        fd_dsp = ast.literal_eval(row1.files_down)
                        #                        print fd_dsp
                        fdp_tmp.append(bubble_sort.sort(fd_dsp.keys()))
                        #                        fdp_tmp.append(fd_dsp)
                        worksheet_d.write(rowc, str(fd_dsp.keys()))
                else:
                    pass
                    # worksheet_d.write(rowc, "")
                tmp += 1
            d_date = bubble_sort.sort(d_date)
            for dc in d_date:
                qc = db.session.query(changes).filter(changes.date_ini == dc.replace("-", "/")).all()
                qcp_tmp.append(qc)
                qi = db.session.query(incidents).filter(incidents.datecab == dc).all()
                qip_tmp.append(qi)
                qddp = db.session.query(downtime_logbook).filter(downtime_logbook.date_downtime == dc).all()
                tmp_t = []
                for k in qddp:
                    if re.search(row.hostname, k.file_name):
                        tmp_t.append(k.time_downtime)
                g = str(sum(tmp_t)) + "  min " + " el " + str(dc.split("-")[0]) + "-" + str(
                    months[int(dc.split("-")[1])])
                qddp_tmp.append(g)
                tmp_d.append(str(dc.split("-")[0]) + "-" + str(months[int(dc.split("-")[1])]))

                for qc_aux in qcp_tmp:
                    for c in qc_aux:
                        t1 = datetime.datetime.strptime(c.time_ini, "%H:%M")
                        t2 = datetime.datetime.strptime(c.time_end, "%H:%M")

                        for server in ast.literal_eval(c.listing):
                            cr_list = {}
                            if re.match(row.hostname, server):
                                try:

                                    qdl = db.session.query(downtime_logbook).filter(
                                        between(downtime_logbook.date_downtime, c.date_ini.replace("/", "-"),
                                                c.date_end.replace("/", "-"))).all()

                                    for row_data in qdl:

                                        tc_ini = datetime.datetime.strptime(c.time_ini, "%H:%M")
                                        tc_end = datetime.datetime.strptime(c.time_end, "%H:%M")
                                        td = datetime.datetime.strptime(row_data.time_file, "%H:%M")
                                        date_down = datetime.datetime.strptime(row_data.date_dowtime, "%d-%m-%Y")
                                        aux_date_ini = datetime.datetime.strptime(c.date_ini.replace("/", "-"),
                                                                                  "%d-%m-%Y")
                                        aux_date_end = datetime.datetime.strptime(c.date_end.replace("/", "-"),
                                                                                  "%d-%m-%Y")
                                        if tc_ini <= td <= tc_end and aux_date_ini <= date_down <= aux_date_end:
                                            if not cr_list.has_key(str(c.reference)):
                                                descr = str(dc.split("-")[0]) + "-" + str(
                                                    months[int(dc.split("-")[1])]) + \
                                                        "/Ejecución de RFC  " + str(c.reference) + "/" + str(
                                                    c.description) + "  De  " + tc_ini.strftime(
                                                    "%I:%M %p") + "  a  " + tc_end.strftime("%I:%M %p")
                                                comments_pc.append(descr)
                                                cr_list[c.reference] = ""
                                except:
                                    comments_pc.append("")
                                    #
                                    # for qi_aux in qip_tmp:
                                    #     for i in qi_aux:
                                    #         if row.hostname in i.listing:
                                    #             t1 = datetime.datetime.strptime(i.time_ini, "%H:%M")
                                    #             t2 = datetime.datetime.strptime(i.time_end, "%H:%M")
                                    #             time_aux = t2 - t1
                                    #             time_i.append((time_aux.seconds / 60))
                                    #            comments_pi.append(i.notes)

            # ttcip = sum(time_c) + sum(time_i)
            ttcip = sum(down_total)
            dtp = sum(down_total)
            fdp = '\n'.join(map(str, fdp_tmp)).replace('[', '').replace(']', '').replace('\'', '').replace('u', '')
            ddp = '\n'.join(map(str, qddp_tmp))
            commentsp = '\n'.join(map(str, comments_pc)) + "\n" + '\n'.join(map(str, comments_pi)) + "\n"
            chostname = "H" + str(aux)
            cenvironment = "G" + str(aux)
            cstatus = "I" + str(aux)
            cip = "J" + str(aux)
            cplatform = "K" + str(aux)
            gservices = "L" + str(aux)
            cservices = "M" + str(aux)
            ctime_down = "N" + str(aux)
            cmaintenance = "O" + str(aux)
            csl = "P" + str(aux)
            cdates_down = "Q" + str(aux)
            cfiles_down = "R" + str(aux)
            ccoments = "S" + str(aux)
            csl_form = '=($O$1-(' + ctime_down + '-' + cmaintenance + '))/$O$1'
            worksheet.write(cenvironment, u"PRODUCCION", format_inventario_content)
            worksheet.write(chostname, row.hostname, format_inventario_content)
            worksheet.write(cstatus, str(row.status.encode('utf8')).upper(), format_inventario_content)
            worksheet.write(cip, row.ip_prod, format_inventario_content)
            worksheet.write(cplatform, row.platform, format_inventario_content)
            worksheet.write(gservices, row.service_group, format_inventario_content)
            worksheet.write(cservices, row.service, format_inventario_content)
            format_inventario_content.set_num_format('#,##0.000')
            worksheet.write_formula(csl, csl_form, format_inventario_content)
            worksheet.write(ctime_down, dtp, format_inventario_content)
            worksheet.write(cfiles_down, str(fdp), format_inventario_content)
            if len(str(fdp)) != 0:
                worksheet.write(cdates_down, str(ddp), format_inventario_content)
            else:
                worksheet.write(cdates_down, "", format_inventario_content)

            worksheet.write_formula(csl, csl_form, format_inventario_content)
            if len(commentsp.strip()) != 0:
                worksheet.write(cmaintenance, str(ttcip), format_inventario_content)
                worksheet.write(ccoments, str(commentsp), format_inventario_content)
            else:
                worksheet.write(cmaintenance, 0, format_inventario_content)
                comments2 = '\n'.join(map(str, tmp_d))
                worksheet.write(ccoments, str(comments2), format_inventario_content)
            aux += 1

        copera = "H" + str(aux)
        worksheet.write(copera, "TOTAL", format_inventario)
        ctopera = "I" + str(aux)
        fopera = '=COUNTIF(I8:' + cstatus.strip() + ', \"OPERATIVO\")'
        worksheet.write_formula(ctopera, fopera, format_inventario_content)
        worksheet.write_formula("A15", "=" + ctopera, format_normal1)
        #        calta = "H" + str(aux + 1)
        #        worksheet.write(calta, "ALTA", format_inventario)
        #        ctalta = "I" + str(aux + 1)
        #        ftalta = '=COUNTIF(I8:' + cstatus.strip() + ', \"ALTA\")'
        #        worksheet.write_formula(ctalta, ftalta, format_inventario_content)
        # ctotal = "H" + str(aux + 2)
        # cttotal = "I" + str(aux + 2)
        ctime_downtotal = "N" + str(aux)
        ctime_just = "O" + str(aux)
        # worksheet.write_formula('A15',"=" + ctopera , format_normal1)
        # worksheet.write(ctotal, "TOTAL", format_inventario)
        # worksheet.write_formula(cttotal, '=SUM(' + ctopera + ":" + ctalta + ')', format_inventario_content)
        # worksheet.write_formula("A15", "=" + cttotal, format_inventario_content)
        worksheet.write_formula(ctime_downtotal, '=SUM(' + 'N8' + ':' + 'N' + str(aux - 1) + ')',
                                format_inventario_content)
        worksheet.write_formula(ctime_just, '=SUM(' + 'O8' + ':' + 'O' + str(aux - 1) + ')', format_inventario_content)
        worksheet.write("N1", "t:", format_header)
        worksheet.write_formula("O1", '=B11', format_normal1)
        worksheet.write("N2", "SLA:", format_header)
        worksheet.write("O2", "99.90%", format_inventario_content1)
        worksheet.write("N3", "Promedio:", format_header)
        format_inventario_content1.set_num_format(0x0a)
        worksheet.write_formula("O3", '=AVERAGE(P8:' + csl + ')', format_inventario_content1)

        # seccion 

        contract = u""" """
        month_total_day = calendar.monthrange(int(year),
                                              int(month1))[1]
        month = months[int(month1)]
        period = u"Perí­odo del 1 de   " + str(month) + u"   al   " \
                 + str(month_total_day) + u"  de  " + str(month) + u"  del  " + str(year)
        worksheet.merge_range('T2:W2', 'Inventario UAT', format_inventario)
        worksheet.merge_range('T5:AE5', contract, format_inventario)
        worksheet.merge_range('T6:AE6', period, format_inventario)
        worksheet.write("U7", "AMBIENTE", format_header2)
        worksheet.write("V7", "HOSTNAME", format_header2)
        worksheet.write("W7", "STATUS", format_header2)
        worksheet.write("X7", "IP", format_header2)
        worksheet.write("Y7", "Plataforma", format_header2)
        worksheet.write("Z7", "Grupo de Servicios", format_header2)
        worksheet.write("AA7", "SERVICIO", format_header2)
        worksheet.write("AB7", "Minutos de Tiempo Fuera", format_header2)
        worksheet.write("AC7", "Minutos de Mantenimiento", format_header2)
        worksheet.write("AD7", "SL_Componente", format_header2)
        worksheet.write("AE7", "Fechas Downtime", format_header2)
        worksheet.write("AF7", "Archivos", format_header2)
        worksheet.write("AG7", "Comentarios ", format_header2)

        q = db.session.query(servers).filter(servers.environment == "uat").order_by(servers.id.asc()).all()
        aux = 8

        for row in q:
            down_total = []
            d_date = []
            time_c = []
            time_i = []
            comments_uc = []
            comments_ui = []
            fdu_tmp = []
            qcu_tmp = []
            qiu_tmp = []
            qddu_tmp = []
            tmp_d = []
            q1 = db.session.query(services_status).filter(
                and_(services_status.servers_id == int(row.id),
                     between(services_status.date, str(month_ini), str(month_end)))).all()

            for row1 in q1:
                rowa = "A" + str(tmp)
                rowb = "B" + str(tmp)
                rowc = "C" + str(tmp)
                rowd = "D" + str(tmp)
                aux_date = datetime.datetime.strptime(row1.date, "%d-%m-%Y")
                if aux_mini <= aux_date <= aux_mend:
                    worksheet_d.write(rowa, row1.date)
                    worksheet_d.write(rowb, row1.time_down)
                    worksheet_d.write(rowd, row1.hostname)

                if len(row1.files_down) != 0:
                    if aux_mini <= aux_date <= aux_mend:
                        down_total.append(int(row1.time_down))
                        d_date.append(str(row1.date))
                        fd_dsu = ast.literal_eval(row1.files_down)
                        fdu_tmp.append(bubble_sort.sort(fd_dsu.keys()))
                        #                        fdu_tmp.append(fd_dsu)
                        worksheet_d.write(rowc, str(fd_dsu.keys()))
                else:
                    worksheet_d.write(rowc, "")
                tmp += 1
            d_date = bubble_sort.sort(d_date)
            for dc in d_date:
                qc = db.session.query(changes).filter(changes.date_ini == dc.replace("-", "/")).all()
                qcu_tmp.append(qc)
                qi = db.session.query(incidents).filter(incidents.datecab == dc).all()
                qiu_tmp.append(qi)
                qddu = db.session.query(downtime_logbook).filter(downtime_logbook.date_downtime == dc).all()
                tmp_t = []
                for k in qddu:
                    if re.search(row.hostname, k.file_name):
                        tmp_t.append(k.time_downtime)
                g = str(sum(tmp_t)) + "  min " + " el " + str(dc.split("-")[0]) + "-" + str(
                    months[int(dc.split("-")[1])])
                qddu_tmp.append(g)
                tmp_d.append(str(dc.split("-")[0]) + "-" + str(months[int(dc.split("-")[1])]))

                for qc_aux in qcu_tmp:
                    for c in qc_aux:
                        for server in ast.literal_eval(c.listing):
                            cr_uatlist = {}
                            if re.match(row.hostname, server):
                                try:
                                    qdl = db.session.query(downtime_logbook).filter(
                                        between(downtime_logbook.date_downtime,
                                                c.date_ini.replace("/", "-"),
                                                c.date_end.replace("/", "-"))).all()
                                    for row_data in qdl:
                                        tc_ini = datetime.datetime.strptime(c.time_ini, "%H:%M")
                                        tc_end = datetime.datetime.strptime(c.time_end, "%H:%M")
                                        td = datetime.datetime.strptime(row_data.time_file, "%H:%M")
                                        date_down = datetime.datetime.strptime(row_data.date_dowtime, "%d-%m-%Y")
                                        aux_date_ini = datetime.datetime.strptime(c.date_ini.replace("/", "-"),
                                                                                  "%d-%m-%Y")
                                        aux_date_end = datetime.datetime.strptime(c.date_end.replace("/", "-"),
                                                                                  "%d-%m-%Y")

                                        if tc_ini <= td <= tc_end and aux_date_ini <= date_down <= aux_date_end:
                                            if not cr_uatlist.has_key(str(c.reference)):
                                                descru = str(dc.split("-")[0]) + "-" + str(
                                                    months[int(dc.split("-")[1])]) + \
                                                         "/Ejecución de RFC  " + str(c.reference) + "/" + str(
                                                    c.description) + "  De  " + tc_ini.strftime(
                                                    "%I:%M %p") + "  a  " + tc_end.strftime("%I:%M %p")
                                                comments_uc.append(descru)
                                                cr_uatlist[c.reference] = ""
                                except:
                                    comments_uc.append("")
                                    #
                                    # for qi_aux in qiu_tmp:
                                    #     for i in qi_aux:
                                    #         if row.hostname in i.listing:
                                    #             t1 = datetime.datetime.strptime(i.time_ini, "%H:%M")
                                    #             t2 = datetime.datetime.strptime(i.time_end, "%H:%M")
                                    #             time_aux = t2 - t1
                                    #             time_i.append((time_aux.seconds / 60))
                                    #             comments_ui.append(i.notes)

            # ttciu = sum(time_c) + sum(time_i)
            ttciu = sum(down_total)
            dtu = sum(down_total)
            fdu = '\n'.join(map(str, fdu_tmp)).replace('[', '').replace(']', '').replace('\'', '').replace('u', '')
            ddu = '\n'.join(map(str, qddu_tmp))
            commentsu = '\n'.join(map(str, comments_uc)) + "\n" + '\n'.join(map(str, comments_ui)) + "\n"
            chostname = "V" + str(aux)
            cenvironment = "U" + str(aux)
            cstatus = "W" + str(aux)
            cip = "X" + str(aux)
            cplatform = "Y" + str(aux)
            gservices = "Z" + str(aux)
            cservices = "AA" + str(aux)
            ctime_down = "AB" + str(aux)
            cmaintenance = "AC" + str(aux)
            csl = "AD" + str(aux)
            cdates_down = "AE" + str(aux)
            cfiles_down = "AF" + str(aux)
            ccoments = "AG" + str(aux)
            csl_form = '=($O$1-(' + ctime_down + '-' + cmaintenance + '))/$O$1'
            worksheet.write(cenvironment, "UAT", format_inventario_content)
            worksheet.write(chostname, row.hostname, format_inventario_content)
            worksheet.write(cstatus, str(row.status).upper(), format_inventario_content)
            worksheet.write(cip, row.ip_prod, format_inventario_content)
            worksheet.write(cplatform, row.platform, format_inventario_content)
            worksheet.write(gservices, row.service_group, format_inventario_content)
            worksheet.write(cservices, row.service, format_inventario_content)
            format_inventario_content.set_num_format('#,##0.000')
            worksheet.write_formula(csl, csl_form, format_inventario_content)
            worksheet.write(ctime_down, dtu, format_inventario_content)
            worksheet.write(cfiles_down, str(fdu), format_inventario_content)
            if len(str(fdu)) != 0 or fdu != "":
                worksheet.write(cdates_down, str(ddu), format_inventario_content)
            else:
                worksheet.write(cdates_down, "", format_inventario_content)

            if len(str(commentsu).strip()) != 0:
                worksheet.write(cmaintenance, str(ttciu), format_inventario_content)
                worksheet.write(ccoments, str(commentsu), format_inventario_content)
            else:
                worksheet.write(cmaintenance, 0, format_inventario_content)
                comments2 = '\n'.join(map(str, tmp_d))
                worksheet.write(ccoments, str(comments2), format_inventario_content)
            aux += 1

        copera = "V" + str(aux)
        worksheet.write(copera, "TOTAL", format_inventario)
        ctopera = "W" + str(aux)
        fopera = '=COUNTIF(V8:' + cstatus.strip() + ', \"OPERATIVO\")'
        worksheet.write_formula(ctopera, fopera, format_inventario_content)
        worksheet.write_formula('A16', "=" + ctopera, format_normal1)
        # calta = "U" + str(aux + 1)
        #        worksheet.write(calta, "ALTA", format_inventario)
        #        ctalta = "V" + str(aux + 1)
        #        ftalta = '=COUNTIF(V8:' + cstatus.strip() + ', \"ALTA\")'
        #        worksheet.write_formula(ctalta, ftalta, format_inventario_content)
        #        ctotal = "U" + str(aux + 2)
        #        cttotal = "V" + str(aux + 2)
        ctime_downtotal = "AB" + str(aux)
        ctime_just = "AC" + str(aux)
        #        worksheet.write(ctotal, "TOTAL", format_inventario)
        #        worksheet.write_formula(cttotal, 'SUM(' + ctopera + ":" + ctalta + ')', format_inventario_content)
        #        worksheet.write_formula("A16", "=" + cttotal, format_inventario_content)
        worksheet.write_formula(ctime_downtotal, "=SUM(AB8" + ":" + ("AB" + str(aux - 1)) + ")",
                                format_inventario_content)
        worksheet.write_formula(ctime_just, "=SUM(AC8" + ":" + ("AC" + str(aux - 1)) + ")", format_inventario_content)
        worksheet.write("AA1", "t:", format_header)
        worksheet.write_formula("AB1", '=B11', format_normal1)
        worksheet.write("AA2", "SLA:", format_header)
        worksheet.write("AB2", "95.00%", format_inventario_content1)
        worksheet.write("AA3", "Promedio:", format_header)
        format_inventario_content1.set_num_format(0x0a)
        worksheet.write_formula("AB3", '=AVERAGE(AD8:' + csl + ")", format_inventario_content1)

        # seccion 4 

        contract = u"""D"""
        month_total_day = calendar.monthrange(int(year), int(month1))[1]
        month = months[int(month1)]
        period = u"Perí­odo del 1 de   " + str(month) + u"   al   " \
                 + str(month_total_day) + u"  de  " + str(month) + u"  del  " + str(year)
        worksheet.merge_range('AG2:AJ2', 'Inventario Stress', format_inventario)
        worksheet.merge_range('AG5:AR5', contract, format_inventario)
        worksheet.merge_range('AG6:AR6', period, format_inventario)
        worksheet.write("AJ7", "AMBIENTE", format_header2)
        worksheet.write("AK7", "HOSTNAME", format_header2)
        worksheet.write("AL7", "STATUS", format_header2)
        worksheet.write("AM7", "IP", format_header2)
        worksheet.write("AN7", "Plataforma", format_header2)
        worksheet.write("AO7", "Grupo de Servicios", format_header2)
        worksheet.write("AP7", "SERVICIO", format_header2)
        worksheet.write("AQ7", "Minutos de Tiempo Fuera", format_header2)
        worksheet.write("AR7", "Minutos de Mantenimiento", format_header2)
        worksheet.write("AS7", "SL_Componente", format_header2)
        worksheet.write("AT7", "Fechas Downtime", format_header2)
        worksheet.write("AU7", "Archivos", format_header2)
        worksheet.write("AV7", "Comentarios", format_header2)

        q = db.session.query(servers).filter(servers.environment == "stress").order_by(servers.id.asc()).all()
        aux = 8

        for row in q:
            down_total = []
            d_date = []
            # time_c = []
            # time_i = []
            comments_sc = []
            comments_si = []
            fds_tmp = []
            qcs_tmp = []
            qis_tmp = []
            qddd_tmp = []
            tmp_d = []
            q1 = db.session.query(services_status).filter(
                and_(services_status.hostname == row.hostname,
                     between(services_status.date, str(month_ini), str(month_end)))).all()
            for row1 in q1:
                rowa = "A" + str(tmp)
                rowb = "B" + str(tmp)
                rowc = "C" + str(tmp)
                rowd = "D" + str(tmp)
                aux_date = datetime.datetime.strptime(row1.date, "%d-%m-%Y")
                if aux_mini <= aux_date <= aux_mend:
                    worksheet_d.write(rowb, row1.time_down)
                    worksheet_d.write(rowd, row1.hostname)

                    if len(row1.files_down) != 0:
                        if aux_mini <= aux_date <= aux_mend:
                            down_total.append(int(row1.time_down))
                            d_date.append(str(row1.date))
                            worksheet_d.write(rowa, row1.date)
                            fd_dss = ast.literal_eval(row1.files_down)

                            fds_tmp.append(bubble_sort.sort(fd_dss.keys()))
                            #                        fds_tmp.append(fd_dss)
                            worksheet_d.write(rowc, str(fd_dss.keys()))
                    else:
                        worksheet_d.write(rowa, "")
                        worksheet_d.write(rowc, "")
                    tmp += 1
            d_date = bubble_sort.sort(d_date)
            for dc in d_date:
                qc = db.session.query(changes).filter(changes.date_ini == dc.replace("-", "/")).all()
                qcs_tmp.append(qc)
                qi = db.session.query(incidents).filter(incidents.datecab == dc).all()
                qis_tmp.append(qi)
                qdds = db.session.query(downtime_logbook).filter(downtime_logbook.date_downtime == dc).all()
                tmp_t = []
                for k in qdds:
                    if re.search(row.hostname, k.file_name):
                        tmp_t.append(k.time_downtime)
                g = str(sum(tmp_t)) + "  min " + " el " + str(dc.split("-")[0]) + "-" + str(
                    months[int(dc.split("-")[1])])
                qddd_tmp.append(g)
                tmp_d.append(str(dc.split("-")[0]) + "-" + str(months[int(dc.split("-")[1])]))

            for qc_aux in qcs_tmp:
                for c in qc_aux:
                    for server in ast.literal_eval(c.listing):
                        cr_stresslist = {}
                        if re.match(row.hostname, server):
                            try:
                                qdl = db.session.query(downtime_logbook).filter(
                                    between(downtime_logbook.date_downtime, c.date_ini.replace("/", "-"),
                                            c.date_end.replace("/", "-"))).all()

                                for row_data in qdl:
                                    tc_ini = datetime.datetime.strptime(c.time_ini, "%H:%M")
                                    tc_end = datetime.datetime.strptime(c.time_end, "%H:%M")
                                    td = datetime.datetime.strptime(row_data.time_file, "%H:%M")
                                    date_down = datetime.datetime.strptime(row_data.date_dowtime, "%d-%m-%Y")
                                    aux_date_ini = datetime.datetime.strptime(c.date_ini.replace("/", "-"), "%d-%m-%Y")
                                    aux_date_end = datetime.datetime.strptime(c.date_end.replace("/", "-"), "%d-%m-%Y")
                                    if tc_ini <= td <= tc_end and aux_date_ini <= date_down <= aux_date_end:
                                        if not cr_stresslist.has_key(str(c.reference)):
                                            descrs = str(dc.split("-")[0]) + "-" + str(months[int(dc.split("-")[1])]) + \
                                                     "/Ejecución de RFC " + str(c.reference) + "/" + str(
                                                c.description) + "  De  " + tc_ini.strftime(
                                                "%I:%M %p") + "  a  " + tc_end.strftime("%I:%M %p")
                                            comments_sc.append(descrs)
                                            cr_stresslist[str(c.reference)] = ""
                            except:
                                comments_sc.append("")

            # for qi_aux in qis_tmp:
            #     for i in qi_aux:
            #
            #         if row.hostname in i.listing:
            #             t1 = datetime.datetime.strptime(i.time_ini, "%H:%M")
            #             t2 = datetime.datetime.strptime(i.time_end, "%H:%M")
            #             time_aux = t2 - t1
            #             time_i.append((time_aux.seconds / 60))
            #             comments_si.append(i.notes)

            # ttcis = sum(time_c) + sum(time_i)
            ttcis = sum(down_total)
            dts = sum(down_total)
            fds = '\n'.join(map(str, fds_tmp)).replace('[', '').replace(']', '').replace('\'', '').replace('u', '')
            dds = '\n'.join(map(str, qddd_tmp))
            commentss = '\n'.join(map(str, comments_sc)) + "\n" + '\n'.join(map(str, comments_si)) + "\n"
            chostname = "AK" + str(aux)
            cenvironment = "AJ" + str(aux)
            cstatus = "AL" + str(aux)
            cip = "AM" + str(aux)
            cplatform = "AN" + str(aux)
            gservices = "AO" + str(aux)
            cservices = "AP" + str(aux)
            ctime_down = "AQ" + str(aux)
            cmaintenance = "AR" + str(aux)
            csl = "AS" + str(aux)
            cdates_down = "AT" + str(aux)
            cfiles_down = "AU" + str(aux)
            ccoments = "AV" + str(aux)
            csl_form = '=($O$1-(' + ctime_down + '-' + cmaintenance + '))/$O$1'

            worksheet.write(cenvironment, "STRESS", format_inventario_content)
            worksheet.write(chostname, row.hostname, format_inventario_content)
            worksheet.write(cstatus, str(row.status).upper(), format_inventario_content)
            worksheet.write(cip, row.ip_prod, format_inventario_content)
            worksheet.write(cplatform, row.platform, format_inventario_content)
            worksheet.write(gservices, row.service_group, format_inventario_content)
            worksheet.write(cservices, row.service, format_inventario_content)
            format_inventario_content.set_num_format('#,##0.000')
            worksheet.write_formula(csl, csl_form, format_inventario_content)
            worksheet.write(ctime_down, dts, format_inventario_content)
            worksheet.write(cfiles_down, str(fds), format_inventario_content)
            if len(str(fds)) != 0:
                worksheet.write(cdates_down, str(dds), format_inventario_content)
            else:

                worksheet.write(cdates_down, "", format_inventario_content)

            if len(commentss.strip()) != 0:
                worksheet.write(cmaintenance, str(ttcis), format_inventario_content)
                worksheet.write(ccoments, str(commentss), format_inventario_content)
            else:
                worksheet.write(cmaintenance, 0, format_inventario_content)
                comments2 = '\n'.join(map(str, tmp_d))
                worksheet.write(ccoments, str(comments2), format_inventario_content)

            aux += 1

        # copera = "AH" + str(aux)
        #        worksheet.write(copera, "OPERACIONAL", format_inventario)
        calta = "AH" + str(aux + 1)
        #        worksheet.write(calta, "ALTA", format_inventario)
        ctotal = "AH" + str(aux + 2)
        #        worksheet.write(ctotal, "TOTAL", format_inventario)

        copera = "AJ" + str(aux)
        worksheet.write(copera, "TOTAL", format_inventario)
        ctopera = "AK" + str(aux)
        fopera = '=COUNTIF(AL8:' + cstatus.strip() + ', \"OPERATIVO\")'
        worksheet.write_formula(ctopera, fopera, format_inventario_content)
        worksheet.write_formula('A17', "=" + ctopera, format_normal1)
        #        calta = "AH" + str(aux + 1)
        #        worksheet.write(calta, "ALTA", format_inventario)
        #        ctalta = "AI" + str(aux + 1)
        #        ftalta = '=COUNTIF(AI8:' + cstatus.strip() + ', \"ALTA\")'
        #        worksheet.write_formula(ctalta, ftalta, format_inventario_content)
        #        ctotal = "AH" + str(aux + 2)
        #        cttotal = "AI" + str(aux + 2)
        ctime_downtotal = "AQ" + str(aux)
        ctime_just = "AR" + str(aux)
        #        worksheet.write(ctotal, "TOTAL", format_inventario)
        #        worksheet.write_formula(cttotal, 'SUM(' + ctopera + ":" + ctalta + ')', format_inventario_content)
        #        worksheet.write_formula("A17", "=" + cttotal, format_inventario_content)
        worksheet.write_formula(ctime_downtotal, "=SUM(AQ8" + ":" + "AQ" + str(aux - 1) + ")",
                                format_inventario_content)
        worksheet.write_formula(ctime_just, "=SUM(AR8" + ":" + "AR" + str(aux - 1) + ")", format_inventario_content)
        worksheet.write("AN1", "t:", format_header)
        worksheet.write_formula("AO1", '=B11', format_normal1)
        worksheet.write("AN2", "SLA:", format_header)
        worksheet.write("AO2", "95.00%", format_inventario_content1)
        worksheet.write("AN3", "Promedio:", format_header)
        format_inventario_content1.set_num_format(0x0a)
        worksheet.write_formula("AO3", '=AVERAGE(AS8:' + csl + ")", format_inventario_content1)

        workbook.close()
        return excel, excel_name


class DatabaseExport:
    """
    A CSV writer which will write rows to CSV file "f",
    which is encoded in the given encoding.
    """

    def __init__(self):
        pass

    def csv_export(self, bd, servers):
        inventory_file = FileConfig.inventory_file()
        if os.path.exists(inventory_file):
            os.remove(inventory_file)
        q = bd.session.query(servers).order_by(servers.id).all()
        csv_file = open(inventory_file, 'wt')
        writer = csv.writer(csv_file)
        writer.writerow((
            u' AMBIENTE', u' HOSTNAME', u' STATUS', u' IP ADMINISTRATIVA', u'IP PRODUCTIVA', u' PLATAFORMA',
            u' SERVICIO', u' PROVEERDOR'))
        for rows in q:
            writer.writerow((unicode(rows.environment).encode('utf-8'), unicode(rows.hostname).encode('utf-8'),
                             unicode(rows.status).encode('utf-8'), unicode(rows.ip_admin).encode('utf-8'),
                             unicode(rows.ip_prod).encode('utf-8'), unicode(rows.platform).encode('utf-8'),
                             unicode(rows.service).encode('utf-8'),
                             'NOVELL'.encode('utf-8')))
        csv_file.close()


if __name__ == "__main__":
    obj = CreateReport()
    obj.create_excel()
